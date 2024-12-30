# -*- coding: utf-8 -*-
"""
/***************************************************************************
 XSDConverter
                        A tool for converting XSD to other formats
                        -------------------
        begin           : 2024-12-24
        git sha         : $Format:%H$
        copyright       : (C) 2024 by https://github.com/kancve
 ***************************************************************************/

/***************************************************************************
 *                                                                         *
 *   This program is free software; you can redistribute it and/or modify  *
 *   it under the terms of the GNU General Public License as published by  *
 *   the Free Software Foundation; either version 2 of the License, or     *
 *   (at your option) any later version.                                   *
 *                                                                         *
 ***************************************************************************/
"""

from openpyxl.comments import Comment
from openpyxl.styles import Alignment
from openpyxl import Workbook
import glob
import argparse
from zs_convert import ZSConvert


def to_zserio(zs_elements, zs_file, ignorecomments=False, sortchildren=False):
    with open(zs_file, 'w', encoding='utf-8') as zs_file:
        for zs_element in zs_elements:
            zs_file.write(zs_element.text(ignorecomments, sortchildren))


def to_xlsx(zs_elements, xlsx, ignorecomments=False, sortchildren=False):
    od_workbook = Workbook()
    od_sheet = od_workbook.active
    od_sheet.title = "OpenDRIVE"

    root_element = None
    excel_elements = {}
    for zs_element in zs_elements:
        if zs_element.name == od_sheet.title:
            root_element = zs_element
        if sortchildren:
            zs_element.children.sort()
        excel_elements[zs_element.type] = zs_element

    def write_sheet(element, sheet, row=1, column=1) -> None:
        cursor_row = row
        for child in element.children:
            sheet.cell(row=cursor_row, column=column).value = child.name
            sub_child = None
            comment = child.comment or ''
            if child.type in excel_elements:
                sub_child = excel_elements[child.type]
                comment = sub_child.comment or comment
            # Comment
            if not ignorecomments:
                text = f'{child.type}({child.use or ""})\n{comment}'
                sheet.cell(row=cursor_row, column=column).comment = Comment(
                    text=text,
                    author='kancve',
                    # # The width of 420 can display 50 characters, while the height of one line is 20
                    height=20 * (1 + text.count('\n') + len(text) / 50),
                    width=420)
            if sub_child and sub_child.children:
                cursor_row = write_sheet(
                    sub_child, sheet, row=cursor_row, column=column+1)
            else:
                cursor_row += 1

        end_row = cursor_row - 1
        merge_column = column - 1
        if end_row > row and merge_column > 0:
            sheet.merge_cells(
                start_row=row, start_column=merge_column, end_row=end_row, end_column=merge_column)
            sheet.cell(row=row, column=merge_column).alignment = Alignment(
                horizontal='general', vertical='center')

        return cursor_row

    write_sheet(root_element, od_sheet)
    # Set column width
    widths = [23, 44, 40, 20, 29, 22, 27, 17, 20, 20, 12]
    for index, width in enumerate(widths):
        od_sheet.column_dimensions[chr(65 + index)].width = width

    od_workbook.save(xlsx)


parser = argparse.ArgumentParser(
    description='OpenDRIVE xsd to zserio schema / xlsx')
parser.add_argument('--xsddir', type=str, default='xsd_schema-1.8.0', help='xsd file directory')
parser.add_argument('--output', type=str, default='OpenDRIVE-1.8.0', help='output file')
parser.add_argument('--ignorecomments', type=bool, default=False, help='Ignore comments (Default to False)')
parser.add_argument('--sortchildren', type=bool, default=False, help='Sort child nodes or not  (Default to False)')
args = parser.parse_args()

zs_elements = list()

# Load the XSD file
print(f'xsd input directory: {args.xsddir}')
for xsd in glob.glob(f'{args.xsddir}/*.xsd'):
    converter = ZSConvert(xsd)
    zs_elements += converter.to_zserio()

output = args.output + '.zs'
print(f'zs output file: {output}')
to_zserio(zs_elements, output, args.ignorecomments, args.sortchildren)


output = args.output + '.xlsx'
print(f'xlsx output file: {output}')
to_xlsx(zs_elements, output, args.ignorecomments, args.sortchildren)

print('Done\n')
